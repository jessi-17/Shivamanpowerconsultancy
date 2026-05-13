"""
Merge citation-tracker.csv and backlinks-action-plan.csv into a single XLSX file
with two tabs. Run once to generate seo-tracker.xlsx.
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="0052DC", end_color="0052DC", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
SECTION_FONT = Font(bold=True, color="0B1C30", size=12)


def csv_to_sheet(wb: Workbook, csv_path: str, sheet_name: str) -> None:
    """Read a CSV and write it to a new sheet in the workbook."""
    ws = wb.create_sheet(title=sheet_name)
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Style section headers (lines with =====)
                if "═══" in value:
                    cell.fill = SECTION_FILL
                    cell.font = SECTION_FONT
                # Style "SECTION X —" lines
                elif value.startswith("SECTION ") and " — " in value:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                # Style first row of header tables (column headers like #, Priority, etc.)
                elif row_idx > 1 and col_idx == 1 and value in ("#", "Field", "Day", "Metric", "Status", "Avoid"):
                    cell.font = Font(bold=True)
                # Title row
                elif row_idx == 1:
                    cell.font = Font(bold=True, size=13, color="0B1C30")

    # Auto-fit column widths (rough heuristic — max 60 chars)
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            v = row[0]
            if v is not None:
                max_len = max(max_len, min(len(str(v)), 60))
        ws.column_dimensions[col_letter].width = max(max_len + 2, 12)

    # Freeze top 4 rows (title + last-updated + how-to)
    ws.freeze_panes = "A5"


def main() -> None:
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Tab 1: Citation Tracker
    csv_to_sheet(wb, "citation-tracker.csv", "Citation Tracker")

    # Tab 2: Backlinks Action Plan
    csv_to_sheet(wb, "backlinks-action-plan.csv", "Backlinks Action Plan")

    # Add an overview tab at the front
    ws_overview = wb.create_sheet(title="README", index=0)
    overview_rows = [
        ["SHIVA TRAVEL & MANPOWER — SEO TRACKER", ""],
        [""],
        ["Last updated:", "2026-05-13"],
        [""],
        ["This workbook has 2 tabs:"],
        [""],
        ["TAB 1: Citation Tracker",
         "Directory listings — 55+ sites (JustDial, Sulekha, IndiaMART, etc.). Also contains: business master data (Section A), Google Business Profile checklist (Section B), industry memberships (Section D)."],
        [""],
        ["TAB 2: Backlinks Action Plan",
         "Content & PR backlink strategies — press release distribution, local press pitches, Quora, Reddit, YouTube, customer-side content, tracking tools, weekly schedule."],
        [""],
        [""],
        ["WHICH TO WORK ON FIRST"],
        [""],
        ["Week 1:", "Citation Tracker Section C — submit to JustDial (done), Sulekha, IndiaMART, TradeIndia, Yellow Pages India, AskLaila (~3 hours)"],
        ["Week 2:", "Citation Tracker Section B — Google Business Profile completeness (photos, services, posts, Q&A) (~2 hours)"],
        ["Week 3:", "Backlinks Action Plan Section A — Government .gov.in backlinks (eMigrate, MEA, ALMRA verification) (~1 hour)"],
        ["Week 4:", "Backlinks Action Plan Section H — Social profile backlinks (YouTube, IG, FB, LinkedIn) (~30 min)"],
        ["Month 2:", "Backlinks Action Plan Section B — Press release distribution"],
        ["Month 2:", "Backlinks Action Plan Section C — Local press pitches (1 per week)"],
        ["Month 3+:", "Backlinks Action Plan Sections D-G — ongoing content (Quora, Reddit, YouTube, customer)"],
        [""],
        [""],
        ["TARGETS (90 DAYS)"],
        [""],
        ["Metric", "Today", "30 days", "90 days"],
        ["Total backlinks (GSC)", "~10-50", "50-80", "150+"],
        ["Unique linking domains", "~5-20", "25-40", "60-80"],
        [".gov.in / .edu.in backlinks", "0 likely", "1-2", "3-5"],
        ["Press mentions landed", "0", "1", "3-5"],
        ["Domain Rating (Ahrefs)", "5-10", "12-15", "18-22"],
        ["GBP reviews", "~42", "55", "80+"],
        ["GBP photos", "<10", "25+", "40+"],
        [""],
        [""],
        ["FILES IN PROJECT"],
        [""],
        ["citation-tracker.csv", "Source CSV for Tab 1 (auto-imported)"],
        ["backlinks-action-plan.csv", "Source CSV for Tab 2 (auto-imported)"],
        ["seo-tracker.xlsx", "This file — merged for Google Sheets / Excel"],
        ["merge_seo_trackers.py", "Re-run if you update the CSVs and want to rebuild the XLSX"],
    ]
    for row_idx, row in enumerate(overview_rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws_overview.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=16, color="0052DC")
            elif value in ("WHICH TO WORK ON FIRST", "TARGETS (90 DAYS)", "FILES IN PROJECT"):
                cell.font = Font(bold=True, size=13, color="0B1C30")
                cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
            elif value.startswith("TAB ") and ":" in value:
                cell.font = Font(bold=True, color="0052DC")

    ws_overview.column_dimensions["A"].width = 30
    ws_overview.column_dimensions["B"].width = 100
    ws_overview.column_dimensions["C"].width = 20
    ws_overview.column_dimensions["D"].width = 20

    out_path = "seo-tracker.xlsx"
    wb.save(out_path)
    print(f"Wrote: {out_path}")
    print(f"  - Tab 1: README")
    print(f"  - Tab 2: Citation Tracker")
    print(f"  - Tab 3: Backlinks Action Plan")


if __name__ == "__main__":
    main()
